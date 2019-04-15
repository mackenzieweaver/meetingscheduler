import requests, bs4, os, openpyxl
from openpyxl import Workbook

abrv = {'enero': 'en',
        'febrero': 'febr',
        'marzo': 'mzo',
        'abril': 'abr',
        'mayo': 'mayo',
        'junio': 'jun',
        'julio': 'jul',
        'agosto': 'ag',
        'septiembre': 'sept',
        'noviembre': 'nov',
        'diciembre': 'dic'}

print('month(s): ', end='')
months = input().split()
print('dates: ', end='')
wbname = input()
dates = wbname.split('-')

base = 'https://www.jw.org/es/publicaciones/guia-actividades-reunion-testigos-jehova/%s-2019-mwb/programa-reunion-' % months[0]
if len(months) > 1:
    ext = dates[0] + abrv[months[0]] + '-' + dates[1] + abrv[months[1]]
else:
    ext = wbname + abrv[months[0]]
url = base + ext + '/'
print(url)

res = requests.get(url)
res.raise_for_status()
soup = bs4.BeautifulSoup(res.text, features="lxml")

# Week and months
elem = soup.select('h1[id="p1"]')
wm = elem[0].getText()
# Weekly bible reading
elem = soup.select('strong')
wbr = elem[0].getText()
# Tesoros de la biblia
elem = soup.select('a[class="pubSym-mwb19"]')
tdlb = elem[0].getText()
# Seamos mejores maestros
smm = []
elem = soup.select('div[id="section3"] strong')
for i in range(len(elem)):
    smm.append(elem[i].getText())
# Nuestra vida cristiana
nvc = []
elem = soup.select('div[id="section4"] strong')
for i in range(len(elem) - 1):
    nvc.append(elem[i].getText())            
# songs
song = []
elem = soup.select('[class="pubSym-sjj"]')
for i in range(len(elem)):
    song.append(elem[i].getText())

x = ''

# Open or create workbook
os.chdir('C:\\Users\\mackn\\Documents\\Spiritual\\Meeting Schedules')
if(os.path.isfile('C:\\Users\\mackn\\Documents\\Spiritual\\Meeting Schedules\\%s.xlsx' % months[0])):
    wb = openpyxl.load_workbook('%s.xlsx' % months[0])
else:
    wb = Workbook()    
# excel sheet
sheets = wb.sheetnames
if dates not in sheets:
    wb.create_sheet(wbname)
else: 
    wb.active.title = wbname
wb.active = wb[wbname]

# TITLE
wb.active['A1'] = 'La Herradura'
wb.active['C1'] = 'Programa para la reunión de entre semana'
# WEEK AND MONTH
wb.active['A4'] = wm
wb.active.merge_cells('A4:B5')
# WEEKLY BIBLE READING
wb.active['C4'] = wbr
wb.active.merge_cells('C4:D5')
wb.active['A7'] = '7:00'
wb.active['B7'] = song[0]
wb.active['A8'] = '7:05'
wb.active['B8'] = 'Palabras de introducción (3 mins. o menos)'
# TESOROS DE LA BIBLIA
wb.active['A9'] = 'TESOROS DE LA BIBLIA'
# wb.active['A9'].fill = greyBlue
wb.active.merge_cells('A9:D9')
wb.active['A10'] = '7:08'
wb.active['B10'] = tdlb + ' (10 mins.)'
wb.active['A11'] = '7:18'
wb.active['B11'] = 'Busquemos perlas escondidas (8 mins.)'
wb.active['A12'] = '7:26'
wb.active['B12'] = 'Lectura de la Biblia (4 mins. o menos)'
# SEAMOS MEJORES MAESTROS
wb.active['A13'] = 'SEAMOS MEJORES MAESTROS'
# wb.active['A13'].fill = yellow
wb.active.merge_cells('A13:D13')
wb.active['A14'] = '7:31'
for i in range(len(smm)):
    wb.active.cell(14+i, 2, smm[i])
# NUESTRA VIDA CRISTIANA
wb.active['A18'] = 'NUESTRA VIDA CRISTIANA'
# wb.active['A18'].fill = 
wb.active.merge_cells('A18:D18')
wb.active['A19'] = '7:47' 
wb.active['B19'] = song[1]
for i in range(len(nvc)):
    wb.active.cell(20+i, 2, nvc[i])
wb.active['B22'] = 'Estudio bíblico de la congregación (30 mins.)'
wb.active['B23'] = 'Repaso de la reunion y adelanto de la próxima (3 mins.)'
wb.active['A24'] = '8:40'
wb.active['B24'] = song[2]

# ASSIGNMENTS
wb.active['F5']  = 'President:'
wb.active['F6']  = 'Consejero de la sala auxiliar:'
wb.active['F7']  = 'Oración:'
wb.active['F5']  = 'President:'
wb.active['E9']  = 'Sala auxiliar'
wb.active['E13'] = 'Sala auxiliar'
wb.active['G9']  = 'Auditorio Principal'
wb.active['G13'] = 'Auditorio Principal'
wb.active['F24'] = 'Oración:'
wb.active['D12'] = 'Estudiante:'
    
# excel sheet
wb.save('%s.xlsx' % months[0])
